"""Point-in-time snapshots — self-contained HTML and Excel (.xlsx) exports.

A *snapshot* freezes a report's current rows (already filtered/grouped in the UI)
into a portable artifact that needs no database:

* HTML — a single file with inline styling and data (no external assets), so it
  opens anywhere and reflects exactly what was on screen, including group subtotals.
* Excel — an ``.xlsx`` workbook (via openpyxl) with the detail rows and, when a
  grouping is supplied, group header and subtotal rows.

Files are written under the user data directory in a ``snapshots`` folder.
"""

from __future__ import annotations

import html
from datetime import datetime
from pathlib import Path
from typing import Any

from dbvisual.app.app_settings import data_dir
from dbvisual.app.report_service import (
    Agg,
    flatten_group_rows,
    group_with_subtotals,
)

__all__ = [
    "render_html",
    "snapshot_dir",
    "write_html_snapshot",
    "write_xlsx_snapshot",
]


def snapshot_dir(dir_override: str | Path | None = None) -> Path:
    """Return (creating if needed) the snapshots folder under the data dir."""
    path = data_dir(dir_override) / "snapshots"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def _safe_name(name: str) -> str:
    keep = "-_. "
    cleaned = "".join(c for c in name if c.isalnum() or c in keep).strip()
    return cleaned.replace(" ", "_") or "snapshot"


# -- HTML -------------------------------------------------------------------

_HTML_STYLE = """
body { font-family: system-ui, Arial, sans-serif; margin: 24px; color: #1f2937; }
h1 { font-size: 20px; margin: 0 0 4px; }
.meta { color: #6b7280; font-size: 12px; margin-bottom: 16px; }
table { border-collapse: collapse; width: 100%; font-size: 13px; }
th, td { border: 1px solid #e5e7eb; padding: 6px 8px; text-align: left; }
th { background: #f3f4f6; }
tr.group td { background: #eef2ff; font-weight: 700; }
tr.subtotal td { background: #eef2f7; font-weight: 700; }
""".strip()


def _cell(value: Any) -> str:
    return "" if value is None else html.escape(str(value))


def render_html(
    title: str,
    fields: list[str],
    rows: list[dict[str, Any]],
    *,
    group_by: list[str] | None = None,
    value_aggs: dict[str, Agg] | None = None,
    sort_by: str = "caption",
    descending: bool = False,
) -> str:
    """Render a self-contained HTML document for ``rows``.

    When ``group_by`` is given, the table is rendered as group header / detail /
    subtotal rows; otherwise a flat table of ``fields`` is produced.
    """
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    head = (
        "<!doctype html><html><head><meta charset='utf-8'>"
        f"<title>{html.escape(title)}</title>"
        f"<style>{_HTML_STYLE}</style></head><body>"
        f"<h1>{html.escape(title)}</h1>"
        f"<div class='meta'>Snapshot generated {generated} - {len(rows)} rows</div>"
    )
    if group_by:
        body = _html_grouped(
            fields, rows, group_by, value_aggs or {}, sort_by, descending
        )
    else:
        body = _html_flat(fields, rows)
    return head + body + "</body></html>"


def _html_flat(fields: list[str], rows: list[dict[str, Any]]) -> str:
    header = "".join(f"<th>{html.escape(f)}</th>" for f in fields)
    body = []
    for r in rows:
        cells = "".join(f"<td>{_cell(r.get(f))}</td>" for f in fields)
        body.append(f"<tr>{cells}</tr>")
    return (
        f"<table><thead><tr>{header}</tr></thead><tbody>{''.join(body)}</tbody></table>"
    )


def _html_grouped(
    fields: list[str],
    rows: list[dict[str, Any]],
    group_by: list[str],
    value_aggs: dict[str, Agg],
    sort_by: str,
    descending: bool,
) -> str:
    tree = group_with_subtotals(
        rows, group_by, value_aggs, sort_by=sort_by, descending=descending
    )
    flat = flatten_group_rows(tree, fields)
    cols = ["_group", *fields, "_count"]
    header = "".join(
        f"<th>{html.escape('Group' if c == '_group' else 'Count' if c == '_count' else c)}</th>"
        for c in cols
    )
    body = []
    for r in flat:
        kind = r.get("_type", "detail")
        indent = "\u2003" * int(r.get("_level", 0))
        cells = []
        for c in cols:
            if c == "_group":
                val = indent + str(r.get("_group", "")) if "_group" in r else ""
                cells.append(f"<td>{html.escape(val)}</td>")
            else:
                cells.append(f"<td>{_cell(r.get(c))}</td>")
        body.append(f"<tr class='{kind}'>{''.join(cells)}</tr>")
    return (
        f"<table><thead><tr>{header}</tr></thead><tbody>{''.join(body)}</tbody></table>"
    )


def write_html_snapshot(
    title: str,
    fields: list[str],
    rows: list[dict[str, Any]],
    *,
    group_by: list[str] | None = None,
    value_aggs: dict[str, Agg] | None = None,
    sort_by: str = "caption",
    descending: bool = False,
    dir_override: str | Path | None = None,
) -> Path:
    """Write an HTML snapshot and return its path."""
    doc = render_html(
        title,
        fields,
        rows,
        group_by=group_by,
        value_aggs=value_aggs,
        sort_by=sort_by,
        descending=descending,
    )
    path = snapshot_dir(dir_override) / f"{_safe_name(title)}-{_timestamp()}.html"
    path.write_text(doc, encoding="utf-8")
    return path


# -- Excel ------------------------------------------------------------------


def write_xlsx_snapshot(
    title: str,
    fields: list[str],
    rows: list[dict[str, Any]],
    *,
    group_by: list[str] | None = None,
    value_aggs: dict[str, Agg] | None = None,
    sort_by: str = "caption",
    descending: bool = False,
    dir_override: str | Path | None = None,
) -> Path:
    """Write an ``.xlsx`` snapshot and return its path.

    Without grouping, a single header row + detail rows are written. With grouping,
    group header rows and subtotal rows are interleaved with the detail rows.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Snapshot"

    bold = Font(bold=True)
    group_fill = PatternFill("solid", fgColor="EEF2FF")
    subtotal_fill = PatternFill("solid", fgColor="EEF2F7")

    if group_by:
        header = ["Group", *fields, "Count"]
        ws.append(header)
        for c in ws[1]:
            c.font = bold
        tree = group_with_subtotals(
            rows, group_by, value_aggs or {}, sort_by=sort_by, descending=descending
        )
        for r in flatten_group_rows(tree, fields):
            kind = r.get("_type", "detail")
            indent = "    " * int(r.get("_level", 0))
            group_txt = (indent + str(r.get("_group", ""))) if "_group" in r else ""
            row_values = [group_txt]
            for f in fields:
                row_values.append(r.get(f))
            row_values.append(r.get("_count"))
            ws.append(row_values)
            if kind in ("group", "subtotal"):
                fill = group_fill if kind == "group" else subtotal_fill
                for c in ws[ws.max_row]:
                    c.font = bold
                    c.fill = fill
    else:
        ws.append(list(fields))
        for c in ws[1]:
            c.font = bold
        for r in rows:
            ws.append([r.get(f) for f in fields])

    path = snapshot_dir(dir_override) / f"{_safe_name(title)}-{_timestamp()}.xlsx"
    wb.save(path)
    return path
