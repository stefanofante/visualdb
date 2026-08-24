"""Tests for point-in-time snapshots (self-contained HTML and Excel)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from dbvisual.app.snapshot_service import (
    render_html,
    snapshot_dir,
    write_html_snapshot,
    write_xlsx_snapshot,
)

FIELDS = ["region", "city", "amount"]


def _rows() -> list[dict[str, Any]]:
    return [
        {"region": "North", "city": "Rome", "amount": 100},
        {"region": "North", "city": "Milan", "amount": 40},
        {"region": "South", "city": "Bari", "amount": 200},
    ]


def test_snapshot_dir_created(tmp_path: Path) -> None:
    d = snapshot_dir(tmp_path)
    assert d.exists() and d.is_dir()
    assert d.name == "snapshots"


def test_render_html_flat_contains_data() -> None:
    doc = render_html("Sales", FIELDS, _rows())
    assert "<html" in doc and "</html>" in doc
    assert "Rome" in doc and "Bari" in doc
    assert "200" in doc
    # self-contained: inline style, no external asset references
    assert "<style>" in doc
    assert "http://" not in doc and "https://" not in doc


def test_render_html_grouped_has_subtotals() -> None:
    doc = render_html(
        "Sales",
        FIELDS,
        _rows(),
        group_by=["region"],
        value_aggs={"amount": "sum"},
    )
    assert "class='group'" in doc
    assert "class='subtotal'" in doc
    # North subtotal = 140, South subtotal = 200
    assert "140" in doc and "200" in doc


def test_write_html_snapshot_roundtrip(tmp_path: Path) -> None:
    path = write_html_snapshot("My Report", FIELDS, _rows(), dir_override=tmp_path)
    assert path.exists()
    assert path.suffix == ".html"
    text = path.read_text(encoding="utf-8")
    assert "Milan" in text
    assert "My Report" in text


def test_write_xlsx_flat_rows(tmp_path: Path) -> None:
    path = write_xlsx_snapshot("Flat", FIELDS, _rows(), dir_override=tmp_path)
    assert path.exists() and path.suffix == ".xlsx"
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == FIELDS
    # header + 3 detail rows
    assert ws.max_row == 4
    values = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]
    assert ("North", "Rome", 100) in values


def test_write_xlsx_grouped_has_subtotals(tmp_path: Path) -> None:
    path = write_xlsx_snapshot(
        "Grouped",
        FIELDS,
        _rows(),
        group_by=["region"],
        value_aggs={"amount": "sum"},
        dir_override=tmp_path,
    )
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["Group", *FIELDS, "Count"]
    # collect the amount column across all rows and look for the 140 subtotal
    amounts = {row[3].value for row in ws.iter_rows(min_row=2)}
    assert 140.0 in amounts  # North subtotal
    assert 200.0 in amounts  # South subtotal
    # a group header row carries the group caption in the first column
    groups = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert any(g and "region: North" in g for g in groups)
